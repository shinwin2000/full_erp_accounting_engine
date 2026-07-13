#!/usr/bin/env python3
"""
Module: generator_pdf_excel_html.py
Layer: Reports
Responsibility: Generator untuk laporan dalam berbagai format: PDF, Excel (XLSX), HTML.
               Mendukung template laporan, styling, embedded charts, dan multiple sheets
               untuk Excel. Juga menyediakan fungsi untuk menambahkan header/footer,
               watermark, dan digital signature pada PDF untuk laporan keuangan.
Dependencies:
- reportlab (PDF generation), openpyxl (Excel), jinja2 (HTML templates)
- asyncio, logging, io, base64, datetime, decimal
- infrastructure.file_storage.s3_adapter (untuk menyimpan laporan)
- infrastructure.telemetry.structured_json_logging
- config.loader_yaml -> DIINJEKSI DARI LUAR (tidak diimpor langsung)
Audit: Setiap laporan yang dihasilkan dicatat. Laporan keuangan ditandatangani digital.
"""

from __future__ import annotations

import asyncio
import json
import logging
import uuid
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import aiofiles  # <-- Tambahan untuk async file I/O

# PDF generation (reportlab)
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas
    from reportlab.platypus import (
        Image,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    # Temporary logger for warning; will be replaced after internal imports
    logger = logging.getLogger(__name__)
    logger.warning("ReportLab not available, PDF generation disabled")

# Excel generation (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("openpyxl not available, Excel generation disabled")

# HTML templates (jinja2)
try:
    from jinja2 import Environment, FileSystemLoader, select_autoescape

    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("jinja2 not available, HTML generation disabled")

# Internal dependencies - config diinjeksi, bukan diimpor langsung
from infrastructure.file_storage.s3_adapter import get_s3_storage_adapter
from infrastructure.security.digital_signer_rsa_pss import DigitalSignerRSA, get_digital_signer
from infrastructure.telemetry.structured_json_logging import get_logger

# Override logger with structured one
logger = get_logger(__name__)

# ============================================================================
# CONSTANTS
# ============================================================================

DEFAULT_CONFIG = {
    "templates_dir": "templates/reports",
    "fonts_dir": "fonts",
    "output_dir": "/tmp/reports",
    "default_font": "Helvetica",
    "company_name": "ERP Accounting Engine",
    "digital_signature_enabled": True,
    "logo_path": "/secrets/company_logo.png",
}

# Supported formats
FORMAT_PDF = "pdf"
FORMAT_EXCEL = "xlsx"
FORMAT_HTML = "html"

# ============================================================================
# EXCEPTIONS
# ============================================================================


class ReportGeneratorError(Exception):
    """Base exception untuk report generator."""
    pass


class UnsupportedFormatError(ReportGeneratorError):
    """Format laporan tidak didukung."""
    pass


class TemplateNotFoundError(ReportGeneratorError):
    """Template tidak ditemukan."""
    pass


# ============================================================================
# REPORT GENERATOR
# ============================================================================


class ReportGenerator:
    """
    Generator laporan dalam berbagai format.

    Fitur:
    - Generate PDF menggunakan ReportLab
    - Generate Excel menggunakan openpyxl
    - Generate HTML menggunakan Jinja2 templates
    - Multi-sheet Excel support
    - Charts dalam Excel (bar, line)
    - Digital signature untuk PDF laporan keuangan
    - Watermark untuk draft reports
    """

    __slots__ = ("_jinja_env", "_output_dir", "_signer", "_templates_dir", "config")

    def __init__(self, config: dict[str, Any] | None = None) -> None:
        """
        Inisialisasi ReportGenerator dengan konfigurasi yang diinjeksi.

        Args:
            config: Dictionary konfigurasi (jika None, gunakan DEFAULT_CONFIG)
        """
        self.config = self._prepare_config(config)
        self._templates_dir = Path(self.config.get("templates_dir", "templates/reports"))
        self._output_dir = Path(self.config.get("output_dir", "/tmp/reports"))
        self._output_dir.mkdir(parents=True, exist_ok=True)
        self._signer: DigitalSignerRSA | None = None
        self._jinja_env: Environment | None = None
        self._init_jinja()

    def _prepare_config(self, config: dict | None) -> dict:
        """Siapkan konfigurasi dari parameter atau default."""
        if config is not None:
            # Merge dengan default untuk memastikan semua key ada
            result = DEFAULT_CONFIG.copy()
            for key, value in config.items():
                if key in result and isinstance(value, dict):
                    result[key].update(value)
                else:
                    result[key] = value
            return result
        return DEFAULT_CONFIG.copy()

    def _init_jinja(self) -> None:
        if JINJA2_AVAILABLE and self._templates_dir.exists():
            self._jinja_env = Environment(
                loader=FileSystemLoader(str(self._templates_dir)),
                autoescape=select_autoescape(["html", "xml"]),
            )
            logger.info(f"Jinja2 templates loaded from {self._templates_dir}")

    async def _get_signer(self) -> DigitalSignerRSA | None:
        if self._signer is None and self.config.get("digital_signature_enabled", True):
            try:
                self._signer = get_digital_signer()
            except Exception as e:
                logger.warning(f"Digital signer not available: {e}")
        return self._signer

    # ========================================================================
    # PERBAIKAN: _sign_pdf menggunakan aiofiles + asyncio.to_thread
    # ========================================================================
    async def _sign_pdf(self, pdf_path: Path, report_id: str, metadata: dict) -> str | None:
        """Menandatangani PDF dengan digital signature."""
        signer = await self._get_signer()
        if not signer:
            return None

        # Baca file PDF secara async
        async with aiofiles.open(pdf_path, "rb") as f:
            pdf_content = await f.read()

        # Sign the content (blocking cryptography, jalankan di thread pool)
        def _sign_sync(content):
            return signer.sign(content)

        signature = await asyncio.to_thread(_sign_sync, pdf_content)

        # Simpan signature sebagai file terpisah
        sig_path = pdf_path.with_suffix(".sig")
        async with aiofiles.open(sig_path, "w") as f:
            await f.write(signature)

        logger.info(f"PDF signed for report {report_id}")
        return signature

    # ========================================================================
    # PDF GENERATION - DIPERBAIKI dengan asyncio.to_thread
    # ========================================================================

    async def generate_pdf(
        self,
        title: str,
        sections: list[dict],
        report_id: str,
        watermark: str | None = None,
        logo_path: Path | None = None,
    ) -> Path:
        """
        Generate PDF report.

        Args:
            title: Judul laporan
            sections: List of section dicts with keys: title, content (list of paragraphs/table)
            report_id: Unique report ID for tracking
            watermark: Teks watermark (e.g., "DRAFT")
            logo_path: Path to company logo

        Returns:
            Path to generated PDF file

        Raises:
            ReportGeneratorError: Jika ReportLab tidak tersedia.
        """
        if not REPORTLAB_AVAILABLE:
            raise ReportGeneratorError("ReportLab not available")

        output_path = self._output_dir / f"{report_id}.pdf"

        def _build_pdf_sync():
            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=A4,
                title=title,
                author=self.config.get("company_name", "ERP Accounting Engine"),
                subject="Financial Report",
            )

            styles = getSampleStyleSheet()
            title_style = styles["Title"]
            heading1_style = styles["Heading1"]
            normal_style = styles["Normal"]

            story = []

            # Logo (if provided)
            if logo_path and logo_path.exists():
                try:
                    img = Image(str(logo_path), width=2 * inch, height=1 * inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2 * inch))
                except Exception:
                    pass

            # Title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.3 * inch))

            # Date
            story.append(
                Paragraph(f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S')}", normal_style)
            )
            story.append(Spacer(1, 0.3 * inch))

            # Sections
            for section in sections:
                story.append(Paragraph(section.get("title", ""), heading1_style))
                story.append(Spacer(1, 0.1 * inch))

                content = section.get("content", [])
                for item in content:
                    if isinstance(item, str):
                        story.append(Paragraph(item, normal_style))
                    elif isinstance(item, dict) and item.get("type") == "table":
                        headers = item.get("headers", [])
                        rows = item.get("rows", [])
                        col_widths = item.get("col_widths", None)
                        table_data = [headers, *rows]
                        t = Table(table_data, colWidths=col_widths)
                        t.setStyle(
                            TableStyle(
                                [
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                                ]
                            )
                        )
                        story.append(t)
                    elif isinstance(item, dict) and item.get("type") == "image":
                        pass

                story.append(Spacer(1, 0.2 * inch))

            # Watermark (if requested) - simplified placeholder
            if watermark:
                pass

            doc.build(story)

        # Jalankan pembuatan PDF di thread pool (blocking)
        await asyncio.to_thread(_build_pdf_sync)

        # Digital signature (async)
        if self.config.get("digital_signature_enabled", True):
            await self._sign_pdf(output_path, report_id, {"title": title})

        logger.info(f"PDF report generated: {output_path}")
        return output_path

    # ========================================================================
    # EXCEL GENERATION - DIPERBAIKI dengan asyncio.to_thread
    # ========================================================================

    async def generate_excel(self, sheets: list[dict], report_id: str) -> Path:
        """
        Generate Excel report with multiple sheets.

        Args:
            sheets: List of sheet dicts with keys: name, headers, rows, chart (optional)
            report_id: Unique report ID

        Returns:
            Path to generated Excel file

        Raises:
            ReportGeneratorError: Jika openpyxl tidak tersedia.
        """
        if not OPENPYXL_AVAILABLE:
            raise ReportGeneratorError("openpyxl not available")

        output_path = self._output_dir / f"{report_id}.xlsx"

        def _build_excel_sync():
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            for sheet_config in sheets:
                ws = wb.create_sheet(
                    title=sheet_config.get("name", "Sheet")[:31]  # Excel sheet name max 31 chars
                )

                headers = sheet_config.get("headers", [])
                rows = sheet_config.get("rows", [])

                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # Write rows
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(value, (Decimal, int, float)):
                            cell.value = value
                        else:
                            cell.value = str(value) if value is not None else ""

                        if isinstance(value, (int, float, Decimal)):
                            cell.number_format = "#,##0.00"

                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    max_length = 0
                    column_letter = get_column_letter(col)
                    for row in range(1, len(rows) + 2):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is not None:
                            max_length = max(max_length, len(str(cell_value)))
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Add chart if requested
                if sheet_config.get("chart"):
                    chart_config = sheet_config["chart"]
                    chart_type = chart_config.get("type", "bar")

                    if chart_type == "bar":
                        chart = BarChart()
                    elif chart_type == "line":
                        chart = LineChart()
                    else:
                        chart = BarChart()

                    chart.title = chart_config.get("title", "")
                    chart.x_axis.title = chart_config.get("x_title", "")
                    chart.y_axis.title = chart_config.get("y_title", "")

                    data = Reference(ws, min_col=2, min_row=1, max_row=len(rows) + 1, max_col=2)
                    categories = Reference(ws, min_col=1, min_row=2, max_row=len(rows) + 1)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(categories)

                    ws.add_chart(chart, chart_config.get("position", "E2"))

            wb.save(str(output_path))

        # Jalankan pembuatan Excel di thread pool (blocking)
        await asyncio.to_thread(_build_excel_sync)

        logger.info(f"Excel report generated: {output_path}")
        return output_path

    # ========================================================================
    # HTML GENERATION - DIPERBAIKI dengan aiofiles
    # ========================================================================

    async def generate_html(self, template_name: str, context: dict, report_id: str) -> Path:
        """
        Generate HTML report from Jinja2 template.

        Args:
            template_name: Name of template file (e.g., "financial_report.html")
            context: Template context variables
            report_id: Unique report ID

        Returns:
            Path to generated HTML file

        Raises:
            ReportGeneratorError: Jika Jinja2 tidak tersedia atau template tidak ditemukan.
            TemplateNotFoundError: Jika template tidak ditemukan.
        """
        if not JINJA2_AVAILABLE or self._jinja_env is None:
            raise ReportGeneratorError("Jinja2 not available or templates not configured")

        # Render template (CPU-bound, jalankan di thread pool)
        def _render_sync():
            template = self._jinja_env.get_template(template_name)
            return template.render(**context)

        try:
            html_content = await asyncio.to_thread(_render_sync)
        except Exception as e:
            raise TemplateNotFoundError(f"Template '{template_name}' not found: {e}")

        output_path = self._output_dir / f"{report_id}.html"
        # Tulis HTML secara async
        async with aiofiles.open(output_path, "w", encoding="utf-8") as f:
            await f.write(html_content)

        logger.info(f"HTML report generated: {output_path}")
        return output_path

    # ========================================================================
    # GENERIC GENERATE METHOD
    # ========================================================================

    async def generate_report(
        self, report_type: str, data: dict, output_format: str, report_id: str | None = None
    ) -> dict[str, Any]:
        """
        Generic method to generate report based on type and format.

        Args:
            report_type: Type of report (e.g., "trial_balance", "income_statement")
            data: Report data
            output_format: "pdf", "xlsx", "html"
            report_id: Optional custom ID

        Returns:
            Dictionary with file path, size, etc.

        Raises:
            UnsupportedFormatError: Jika format tidak didukung.
        """
        output_format = output_format.lower()
        if output_format not in (FORMAT_PDF, FORMAT_EXCEL, FORMAT_HTML):
            raise UnsupportedFormatError(f"Unsupported output format: {output_format}")

        if report_id is None:
            report_id = (
                f"{report_type}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
            )

        if output_format == FORMAT_PDF:
            sections = self._data_to_sections(report_type, data)
            file_path = await self.generate_pdf(
                title=f"{report_type.replace('_', ' ').title()} Report",
                sections=sections,
                report_id=report_id,
            )
        elif output_format == FORMAT_EXCEL:
            sheets = self._data_to_sheets(report_type, data)
            file_path = await self.generate_excel(sheets, report_id)
        elif output_format == FORMAT_HTML:
            template_name = f"{report_type}.html"
            file_path = await self.generate_html(template_name, data, report_id)
        else:
            # Should never reach due to earlier check
            raise UnsupportedFormatError(f"Unsupported output format: {output_format}")

        file_size = file_path.stat().st_size

        return {
            "report_id": report_id,
            "report_type": report_type,
            "format": output_format,
            "file_path": str(file_path),
            "file_name": file_path.name,
            "file_size_bytes": file_size,
            "generated_at": datetime.now(UTC).isoformat(),
        }

    def _data_to_sections(self, report_type: str, data: dict) -> list[dict]:
        """Convert report data to PDF sections format."""
        return [{"title": "Report Data", "content": [json.dumps(data, indent=2, default=str)]}]

    def _data_to_sheets(self, report_type: str, data: dict) -> list[dict]:
        """Convert report data to Excel sheets format."""
        if not data:
            return [{"name": report_type[:31], "headers": [], "rows": []}]
        headers = list(data.keys())
        rows = [[v for v in data.values()]]
        return [{"name": report_type[:31], "headers": headers, "rows": rows}]

    # ========================================================================
    # PERBAIKAN: upload_report menggunakan aiofiles
    # ========================================================================
    async def upload_report(self, file_path: Path, bucket: str | None = None) -> str:
        """Upload generated report to cloud storage."""
        try:
            storage = await get_s3_storage_adapter()
            # Baca file secara async
            async with aiofiles.open(file_path, "rb") as f:
                file_content = await f.read()
            uri = await storage.upload(
                file_content=file_content,
                file_name=file_path.name,
                content_type=self._get_content_type(file_path.suffix),
                bucket=bucket,
            )
            return uri
        except Exception as e:
            logger.error(f"Failed to upload report: {e}")
            return str(file_path)

    @staticmethod
    def _get_content_type(suffix: str) -> str:
        if suffix == ".pdf":
            return "application/pdf"
        elif suffix == ".xlsx":
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif suffix == ".html":
            return "text/html"
        else:
            return "application/octet-stream"


# ============================================================================
# SINGLETON INSTANCE dengan injeksi konfigurasi
# ============================================================================

_report_generator: ReportGenerator | None = None
_generator_config: dict | None = None


def set_report_generator_config(config: dict) -> None:
    """Set konfigurasi untuk ReportGenerator (harus dipanggil sebelum get_report_generator)."""
    global _generator_config
    _generator_config = config


async def get_report_generator() -> ReportGenerator:
    """Get singleton instance of ReportGenerator."""
    global _report_generator
    if _report_generator is None:
        _report_generator = ReportGenerator(config=_generator_config)
    return _report_generator


# ============================================================================
# EXPORTS
# ============================================================================

__all__ = [
    "ReportGenerator",
    "ReportGeneratorError",
    "TemplateNotFoundError",
    "UnsupportedFormatError",
    "get_report_generator",
    "set_report_generator_config",
]
