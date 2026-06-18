#!/usr/bin/env python3

"""
Module: umkm_simplified.py

Layer: 8 - Application / Workflows

Responsibility:
    Workflow sederhana untuk UMKM (Usaha Mikro Kecil Menengah).
    Mencakup:
    - Pencatatan transaksi kas masuk dan kas keluar (cash basis)
    - Kategorisasi pendapatan dan beban
    - Perhitungan laba rugi sederhana
    - Perhitungan PPh Final UMKM (0.5% dari omzet)
    - Laporan keuangan sederhana (Laba Rugi, Arus Kas, Neraca sederhana)
    - Export ke format Excel untuk pelaporan

Dependencies:
    - application/service_layer/service_umkm.py (UMKMService)
    - application/service_layer/service_journal.py (JournalService)
    - application/service_layer/service_tax.py (TaxService)
    - application/commands_cqrs/command_bus_unified.py (Command, CommandResult)

Audit:
    Setiap transaksi UMKM dicatat dengan timestamp.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_EVEN, Decimal
from enum import Enum
from typing import TYPE_CHECKING, Any

from application.commands_cqrs.command_bus_unified import Command, CommandResult

if TYPE_CHECKING:
    from uuid import UUID

    from application.service_layer.service_journal import JournalService
    from application.service_layer.service_tax import TaxService
    from application.service_layer.service_umkm import UMKMService
    from kernel.sealed_gate import SealedGate

logger = logging.getLogger(__name__)


class TransactionCategory(Enum):
    INCOME_SALES = "INCOME_SALES"
    INCOME_SERVICE = "INCOME_SERVICE"
    INCOME_OTHER = "INCOME_OTHER"
    EXPENSE_COGS = "EXPENSE_COGS"
    EXPENSE_SALARY = "EXPENSE_SALARY"
    EXPENSE_RENT = "EXPENSE_RENT"
    EXPENSE_UTILITY = "EXPENSE_UTILITY"
    EXPENSE_OTHERS = "EXPENSE_OTHERS"


class UMKMWorkflowCommand(Command):
    """Command untuk workflow UMKM."""

    __slots__ = (
        "action",
        "amount",
        "category",
        "description",
        "export_format",
        "legal_entity_id",
        "payment_method",
        "period_end",
        "period_start",
        "reference_number",
        "transaction_date",
    )

    def __init__(
        self,
        legal_entity_id: UUID,
        action: str,
        transaction_date: date | None = None,
        amount: Decimal | None = None,
        category: str | None = None,
        description: str | None = None,
        payment_method: str = "CASH",
        reference_number: str | None = None,
        period_start: date | None = None,
        period_end: date | None = None,
        export_format: str = "excel",
        user_id: UUID | None = None,
        correlation_id: str | None = None,
    ):
        super().__init__(
            command_type="UMKMWorkflowCommand", user_id=user_id, correlation_id=correlation_id
        )
        self.legal_entity_id = legal_entity_id
        self.action = action
        self.transaction_date = transaction_date
        self.amount = amount
        self.category = category
        self.description = description
        self.payment_method = payment_method
        self.reference_number = reference_number
        self.period_start = period_start
        self.period_end = period_end
        self.export_format = export_format

    def to_dict(self) -> dict[str, Any]:
        data = super().to_dict()
        data.update(
            {
                "legal_entity_id": str(self.legal_entity_id),
                "action": self.action,
                "transaction_date": (
                    self.transaction_date.isoformat() if self.transaction_date else None
                ),
                "amount": float(self.amount) if self.amount else None,
                "category": self.category,
                "description": self.description,
                "payment_method": self.payment_method,
                "reference_number": self.reference_number,
                "period_start": self.period_start.isoformat() if self.period_start else None,
                "period_end": self.period_end.isoformat() if self.period_end else None,
                "export_format": self.export_format,
            }
        )
        return data


class UMKMWorkflowResult:
    def __init__(
        self,
        action: str,
        transaction_id: UUID | None,
        report_data: dict[str, Any] | None,
        file_path: str | None,
        message: str,
    ):
        self.action = action
        self.transaction_id = transaction_id
        self.report_data = report_data
        self.file_path = file_path
        self.message = message


class UMKMWorkflow:
    """
    Workflow sederhana untuk UMKM.
    """

    def __init__(
        self,
        umkm_service: UMKMService,
        journal_service: JournalService,
        tax_service: TaxService,
        sealed_gate: SealedGate | None = None,
    ):
        self._umkm_service = umkm_service
        self._journal_service = journal_service
        self._tax_service = tax_service
        self._sealed_gate = sealed_gate
        self._stats = {"executed": 0, "succeeded": 0, "failed": 0}

    async def execute(self, command: UMKMWorkflowCommand) -> CommandResult:
        self._stats["executed"] += 1

        try:
            if command.action == "RECORD_TRANSACTION":
                result = await self._record_transaction(command)
            elif command.action == "GENERATE_REPORT":
                result = await self._generate_report(command)
            elif command.action == "CALCULATE_TAX":
                result = await self._calculate_tax(command)
            else:
                raise ValueError(f"Unknown action: {command.action}")

            self._stats["succeeded"] += 1
            return CommandResult.success(
                command_id=command.command_id,
                data={
                    "action": result.action,
                    "transaction_id": str(result.transaction_id) if result.transaction_id else None,
                    "file_path": result.file_path,
                    "message": result.message,
                },
            )

        except Exception as e:
            self._stats["failed"] += 1
            logger.exception(f"UMKM workflow failed: {e}")
            return CommandResult.failure(
                command_id=command.command_id, error=str(e), error_code="UMKM_WORKFLOW_ERROR"
            )

    async def _record_transaction(self, command: UMKMWorkflowCommand) -> UMKMWorkflowResult:
        if not command.transaction_date or not command.amount or not command.category:
            raise ValueError("Transaction date, amount, and category required")

        category_enum = None
        try:
            category_enum = TransactionCategory(command.category)
        except ValueError:
            raise ValueError(f"Invalid category: {command.category}")

        is_income = category_enum.value.startswith("INCOME")

        transaction = await self._umkm_service.record_transaction(
            legal_entity_id=command.legal_entity_id,
            transaction_date=command.transaction_date,
            amount=command.amount,
            transaction_type="INCOME" if is_income else "EXPENSE",
            category=command.category,
            description=command.description or "",
            payment_method=command.payment_method,
            reference_number=command.reference_number,
            user_id=command.user_id,
            correlation_id=command.correlation_id,
        )

        await self._umkm_service.create_simple_journal(
            legal_entity_id=command.legal_entity_id,
            transaction_id=transaction.id,
            amount=command.amount,
            transaction_type="INCOME" if is_income else "EXPENSE",
            category=command.category,
            transaction_date=command.transaction_date,
            user_id=command.user_id,
        )

        return UMKMWorkflowResult(
            action="RECORD_TRANSACTION",
            transaction_id=transaction.id,
            report_data=None,
            file_path=None,
            message=f"Transaction recorded: {command.category} of {command.amount}",
        )

    async def _generate_report(self, command: UMKMWorkflowCommand) -> UMKMWorkflowResult:
        if not command.period_start or not command.period_end:
            today = date.today()
            command.period_start = date(today.year, today.month, 1)
            if today.month == 12:
                command.period_end = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                command.period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)

        income_total = await self._umkm_service.get_total_income(
            command.legal_entity_id, command.period_start, command.period_end
        )
        expense_total = await self._umkm_service.get_total_expense(
            command.legal_entity_id, command.period_start, command.period_end
        )
        net_profit = income_total - expense_total

        transactions = await self._umkm_service.list_transactions(
            command.legal_entity_id, command.period_start, command.period_end
        )

        report_data = {
            "period_start": command.period_start.isoformat(),
            "period_end": command.period_end.isoformat(),
            "total_income": float(income_total),
            "total_expense": float(expense_total),
            "net_profit": float(net_profit),
            "transactions": [
                {
                    "date": t.transaction_date.isoformat(),
                    "type": t.transaction_type,
                    "category": t.category,
                    "amount": float(t.amount),
                    "description": t.description,
                }
                for t in transactions
            ],
        }

        file_path = None
        if command.export_format == "excel":
            file_path = await self._export_to_excel(report_data, command)
        elif command.export_format == "csv":
            file_path = await self._export_to_csv(report_data, command)

        return UMKMWorkflowResult(
            action="GENERATE_REPORT",
            transaction_id=None,
            report_data=report_data,
            file_path=file_path,
            message=f"Report generated for {command.period_start} to {command.period_end}",
        )

    async def _calculate_tax(self, command: UMKMWorkflowCommand) -> UMKMWorkflowResult:
        if not command.period_start or not command.period_end:
            today = date.today()
            command.period_start = date(today.year, today.month, 1)
            if today.month == 12:
                command.period_end = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                command.period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)

        total_income = await self._umkm_service.get_total_income(
            command.legal_entity_id, command.period_start, command.period_end
        )

        tax_rate = Decimal("0.005")
        tax_due = (total_income * tax_rate).quantize(Decimal("0"), rounding=ROUND_HALF_EVEN)

        payments = await self._umkm_service.get_tax_payments(
            command.legal_entity_id, command.period_start.year
        )
        total_paid = sum(p.amount for p in payments)
        tax_payable = max(tax_due - total_paid, Decimal("0"))

        tax_data = {
            "period": f"{command.period_start.year}-{command.period_start.month:02d}",
            "gross_revenue": float(total_income),
            "tax_rate": 0.5,
            "tax_due": float(tax_due),
            "tax_paid": float(total_paid),
            "tax_payable": float(tax_payable),
        }

        file_path = None
        if command.export_format == "excel":
            file_path = await self._export_tax_to_excel(tax_data, command)

        return UMKMWorkflowResult(
            action="CALCULATE_TAX",
            transaction_id=None,
            report_data=tax_data,
            file_path=file_path,
            message=f"Tax for {command.period_start.year}-{command.period_start.month:02d}: due {tax_due}, payable {tax_payable}",
        )

    async def _export_to_excel(
        self, report_data: dict[str, Any], command: UMKMWorkflowCommand
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "UMKM Report"

            ws["A1"] = "LAPORATAN UMKM"
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:F1")

            ws["A2"] = f"Periode: {report_data['period_start']} s/d {report_data['period_end']}"
            ws.merge_cells("A2:F2")

            ws["A4"] = "Ringkasan"
            ws["A4"].font = Font(bold=True)
            ws["A5"] = "Total Pendapatan"
            ws["B5"] = report_data["total_income"]
            ws["A6"] = "Total Beban"
            ws["B6"] = report_data["total_expense"]
            ws["A7"] = "Laba Bersih"
            ws["B7"] = report_data["net_profit"]

            ws["A9"] = "Detail Transaksi"
            ws["A9"].font = Font(bold=True)
            headers = ["Tanggal", "Jenis", "Kategori", "Jumlah", "Deskripsi"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            for row_idx, tx in enumerate(report_data["transactions"], start=11):
                ws.cell(row=row_idx, column=1, value=tx["date"])
                ws.cell(row=row_idx, column=2, value=tx["type"])
                ws.cell(row=row_idx, column=3, value=tx["category"])
                ws.cell(row=row_idx, column=4, value=tx["amount"])
                ws.cell(row=row_idx, column=5, value=tx["description"])

            for col in range(1, 6):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            file_path = (
                f"/tmp/umkm_report_{command.legal_entity_id}_{datetime.utcnow().timestamp()}.xlsx"
            )
            wb.save(file_path)
            return file_path

        except ImportError:
            return await self._export_to_csv(report_data, command)

    async def _export_to_csv(
        self, report_data: dict[str, Any], command: UMKMWorkflowCommand
    ) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["UMKM Report"])
        writer.writerow([f"Period: {report_data['period_start']} to {report_data['period_end']}"])
        writer.writerow([])
        writer.writerow(["Summary"])
        writer.writerow(["Total Income", report_data["total_income"]])
        writer.writerow(["Total Expense", report_data["total_expense"]])
        writer.writerow(["Net Profit", report_data["net_profit"]])
        writer.writerow([])
        writer.writerow(["Transactions"])
        writer.writerow(["Date", "Type", "Category", "Amount", "Description"])
        for tx in report_data["transactions"]:
            writer.writerow(
                [tx["date"], tx["type"], tx["category"], tx["amount"], tx["description"]]
            )

        file_path = (
            f"/tmp/umkm_report_{command.legal_entity_id}_{datetime.utcnow().timestamp()}.csv"
        )
        with open(file_path, "w") as f:
            f.write(output.getvalue())
        return file_path

    async def _export_tax_to_excel(
        self, tax_data: dict[str, Any], command: UMKMWorkflowCommand
    ) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["UMKM Tax Report"])
        writer.writerow([f"Period: {tax_data['period']}"])
        writer.writerow([])
        writer.writerow(["Gross Revenue", tax_data["gross_revenue"]])
        writer.writerow(["Tax Rate (%)", tax_data["tax_rate"]])
        writer.writerow(["Tax Due", tax_data["tax_due"]])
        writer.writerow(["Tax Paid", tax_data["tax_paid"]])
        writer.writerow(["Tax Payable", tax_data["tax_payable"]])

        file_path = f"/tmp/umkm_tax_{command.legal_entity_id}_{tax_data['period']}.csv"
        with open(file_path, "w") as f:
            f.write(output.getvalue())
        return file_path

    def get_stats(self) -> dict[str, int]:
        return self._stats


# ============================================================================
# Factory function
# ============================================================================


def create_umkm_workflow(
    umkm_service: UMKMService,
    journal_service: JournalService,
    tax_service: TaxService,
    sealed_gate: SealedGate | None = None,
) -> UMKMWorkflow:
    """Factory untuk membuat workflow UMKM."""
    return UMKMWorkflow(
        umkm_service=umkm_service,
        journal_service=journal_service,
        tax_service=tax_service,
        sealed_gate=sealed_gate,
    )


__all__ = [
    "TransactionCategory",
    "UMKMWorkflow",
    "UMKMWorkflowCommand",
    "UMKMWorkflowResult",
    "create_umkm_workflow",
]
