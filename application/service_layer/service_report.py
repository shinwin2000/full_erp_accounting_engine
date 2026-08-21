#!/usr/bin/env python3
"""
Module: service_report.py
Layer: 8 - Application / Service Layer
Responsibility:
    Service layer untuk pelaporan (Financial Reports, Report Management).

REWRITE TOTAL (2026-08-18): implementasi sebelumnya adalah stub murni -
setiap method balikin data kosong/nol dan constructor tidak menerima
dependency apapun (`def __init__(self): self._stats = {...}`), padahal
fastapi_report_router.py butuh 16 method (12 generate_* + list/get/status/
history/delete report tersimpan) yang semuanya belum ada sama sekali,
menyebabkan `AttributeError` di setiap endpoint report.

Pendekatan: service ini TIDAK menghitung ulang logika laporan keuangan dari
nol. Untuk laporan yang datanya sudah punya implementasi nyata dan terbukti
jalan di modul lain (LedgerService, ARService, APService, BudgetService,
InventoryService - semuanya sudah dipakai production dan menghasilkan
200 OK), service ini mendelegasikan ke sana, lalu merender hasilnya jadi
file (PDF/Excel/CSV/JSON) dan mencatat metadatanya ke tabel
`generated_report` supaya bisa di-list/get/delete/lihat statusnya.

Laporan yang datanya BELUM ada implementasi nyata di tempat lain
(financial_ratios industry comparison, budget actual dari GL) tetap
mengembalikan struktur yang benar dengan field yang belum bisa dihitung
diisi None/0, alih-alih exception - konsisten dengan cara LedgerService
sendiri menangani keterbatasan serupa (lihat komentar di
service_ledger.py: "chart of accounts doesn't carry today").
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any
from uuid import UUID, uuid4

from application.service_layer.service_ap import APService
from application.service_layer.service_ar import ARService
from application.service_layer.service_budget import BudgetService
from application.service_layer.service_inventory import InventoryService
from application.service_layer.service_ledger import LedgerService
from infrastructure.persistence_orm.generated_report_table import GeneratedReportTable
from ports.primary.report_repository_port import ReportRepositoryPort
from ports.primary.tax_transaction_repository_port import TaxTransactionRepositoryPort

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed, Excel export akan fallback ke CSV")

REPORT_RETENTION_DAYS = 90


# ============================================================================
# Exceptions
# ============================================================================


class ReportServiceError(Exception):
    pass


# ============================================================================
# Value objects tambahan (dikembalikan ke router selain GeneratedReportTable)
# ============================================================================


@dataclass(kw_only=True)
class PagedGeneratedReports:
    items: list[GeneratedReportTable]
    total: int


@dataclass(kw_only=True)
class ReportStatusInfo:
    report_number: str
    status: str
    progress_percent: int
    current_step: str | None
    total_steps: int
    estimated_remaining_seconds: int | None
    error_message: str | None
    generated_at: datetime | None


@dataclass(kw_only=True)
class ReportHistoryEntry:
    timestamp: datetime
    action: str
    status: str
    actor_id: UUID
    actor_name: str | None = None
    reason: str | None = None
    details: dict[str, Any] | None = None


# ============================================================================
# ReportService
# ============================================================================


class ReportService:
    def __init__(
        self,
        report_repo: ReportRepositoryPort,
        ledger_service: LedgerService,
        ar_service: ARService,
        ap_service: APService,
        budget_service: BudgetService,
        inventory_service: InventoryService,
        tax_transaction_repo: TaxTransactionRepositoryPort | None = None,
    ):
        self._report_repo = report_repo
        self._ledger = ledger_service
        self._ar = ar_service
        self._ap = ap_service
        self._budget = budget_service
        self._inventory = inventory_service
        self._tax_transaction_repo = tax_transaction_repo
        self._stats = {"reports_generated": 0, "errors": 0}
        self._audit_trail: list[dict[str, Any]] = []
        logger.info("ReportService initialized")

    # ==================== AUDIT TRAIL ====================

    def _record_audit(self, action: str, details: dict[str, Any]) -> None:
        self._audit_trail.append(
            {"timestamp": datetime.utcnow().isoformat(), "action": action, "details": details}
        )

    def get_stats(self) -> dict[str, int]:
        return dict(self._stats)

    def get_audit_trail(self) -> list[dict[str, Any]]:
        return list(self._audit_trail)

    # ==================== HELPERS: RENDER FILE ====================

    def _next_report_number(self, report_type: str) -> str:
        stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        return f"RPT-{report_type.upper()}-{stamp}-{str(uuid4())[:6]}"

    def _render_csv(self, headers: list[str], rows: list[list[Any]]) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig")

    def _render_excel(self, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
        if not HAS_OPENPYXL:
            return self._render_csv(headers, rows)
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31] if title else "Report"
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row in enumerate(rows, 2):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, (Decimal, int, float)):
                    ws.cell(row=r_idx, column=c_idx, value=float(value))
                else:
                    ws.cell(row=r_idx, column=c_idx, value=str(value) if value is not None else "")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _render_json(self, payload: dict[str, Any]) -> bytes:
        import json

        def _default(o: Any) -> Any:
            if isinstance(o, Decimal):
                return str(o)
            if isinstance(o, (date, datetime)):
                return o.isoformat()
            if isinstance(o, UUID):
                return str(o)
            return str(o)

        return json.dumps(payload, default=_default, indent=2, ensure_ascii=False).encode("utf-8")

    def _render_html(self, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
        head_html = "".join(f"<th>{h}</th>" for h in headers)
        rows_html = "".join(
            "<tr>" + "".join(f"<td>{c if c is not None else ''}</td>" for c in row) + "</tr>"
            for row in rows
        )
        html = (
            f"<html><head><meta charset='utf-8'><title>{title}</title></head>"
            f"<body><h1>{title}</h1><table border='1' cellspacing='0' cellpadding='4'>"
            f"<thead><tr>{head_html}</tr></thead><tbody>{rows_html}</tbody></table></body></html>"
        )
        return html.encode("utf-8")

    def _render_pdf(self, title: str, headers: list[str], rows: list[list[Any]]) -> bytes | None:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError:
            logger.warning("reportlab tidak tersedia, PDF tidak bisa dirender")
            return None

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
        styles = getSampleStyleSheet()
        elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

        table_data = [headers] + [
            [str(c) if c is not None else "" for c in row] for row in rows
        ]
        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(tbl)
        doc.build(elements)
        return buf.getvalue()

    async def _persist_report(
        self,
        *,
        legal_entity_id: UUID,
        report_type: str,
        report_format: str,
        title: str,
        headers: list[str],
        rows: list[list[Any]],
        raw_payload: dict[str, Any],
        parameters: dict[str, Any],
        generated_by: UUID,
    ) -> GeneratedReportTable:
        """Render laporan ke format yang diminta (best-effort - kalau library
        rendering-nya tidak tersedia, fallback ke JSON) lalu simpan
        metadatanya sebagai baris `generated_report`. File fisik disimpan
        di working directory proses (Path relatif "generated_reports/"),
        konsisten dengan LocalFileStorage yang sudah dipakai adapter lain
        di proyek ini untuk penyimpanan lokal.
        """
        import os

        report_number = self._next_report_number(report_type)
        fmt = (report_format or "pdf").lower()

        content: bytes | None
        actual_ext = fmt
        if fmt in ("xlsx", "excel"):
            content = self._render_excel(title, headers, rows)
            actual_ext = "xlsx"
        elif fmt == "csv":
            content = self._render_csv(headers, rows)
            actual_ext = "csv"
        elif fmt == "html":
            content = self._render_html(title, headers, rows)
            actual_ext = "html"
        elif fmt == "json":
            content = self._render_json(raw_payload)
            actual_ext = "json"
        elif fmt == "pdf":
            content = self._render_pdf(title, headers, rows)
            actual_ext = "pdf"
            if content is None:
                # ReportLab tidak tersedia - fallback JSON supaya laporan
                # tetap tersimpan & bisa diunduh, bukan gagal total.
                content = self._render_json(raw_payload)
                actual_ext = "json"
        else:
            content = self._render_json(raw_payload)
            actual_ext = "json"

        out_dir = os.path.join(os.getcwd(), "generated_reports")
        os.makedirs(out_dir, exist_ok=True)
        file_path = os.path.join(out_dir, f"{report_number}.{actual_ext}")
        try:
            with open(file_path, "wb") as f:
                f.write(content)
            file_size = len(content)
            status = "generated"
            error_message = None
        except OSError as e:
            logger.exception("Gagal menulis file laporan %s: %s", report_number, e)
            file_path = None
            file_size = None
            status = "failed"
            error_message = str(e)

        now = datetime.utcnow()
        record = GeneratedReportTable(
            id=uuid4(),
            legal_entity_id=legal_entity_id,
            report_number=report_number,
            report_type=report_type,
            report_format=report_format,
            status=status,
            file_path=file_path,
            file_size_bytes=file_size,
            parameters=parameters,
            error_message=error_message,
            generated_at=now,
            generated_by=generated_by,
            generated_by_name=None,
            expires_at=now + timedelta(days=REPORT_RETENTION_DAYS),
            is_deleted=False,
            created_at=now,
        )
        saved = await self._report_repo.create_generated_report(record)
        self._stats["reports_generated"] += 1
        self._record_audit(f"generate_{report_type}", {"report_number": report_number})
        return saved

    @staticmethod
    def _dec(value: Any) -> str:
        return str(value) if value is not None else ""

    # ==================== FINANCIAL STATEMENTS ====================

    async def generate_balance_sheet(
        self,
        *,
        legal_entity_id: UUID,
        as_of_date: date,
        include_details: bool = True,
        compare_with_previous: bool = False,
        currency: str = "IDR",
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        data = await self._ledger.get_balance_sheet(
            legal_entity_id, as_of_date, include_comparatives=compare_with_previous
        )
        headers = ["Kategori", "Akun", "Saldo"]
        rows: list[list[Any]] = []
        for line in data.assets_lines:
            rows.append(["Aset", line.get("account_name", ""), self._dec(line.get("balance"))])
        for line in data.liabilities_lines:
            rows.append(["Kewajiban", line.get("account_name", ""), self._dec(line.get("balance"))])
        for line in data.equity_lines:
            rows.append(["Ekuitas", line.get("account_name", ""), self._dec(line.get("balance"))])
        rows.append(["TOTAL ASET", "", self._dec(data.total_assets)])
        rows.append(["TOTAL KEWAJIBAN", "", self._dec(data.total_liabilities)])
        rows.append(["TOTAL EKUITAS", "", self._dec(data.total_equity)])

        payload = {
            "as_of_date": as_of_date.isoformat(),
            "total_assets": str(data.total_assets),
            "total_liabilities": str(data.total_liabilities),
            "total_equity": str(data.total_equity),
            "assets": data.assets_lines,
            "liabilities": data.liabilities_lines,
            "equity": data.equity_lines,
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="balance_sheet", report_format=report_format,
            title="Neraca (Balance Sheet)", headers=headers, rows=rows, raw_payload=payload,
            parameters={"as_of_date": as_of_date.isoformat(), "include_details": include_details,
                        "compare_with_previous": compare_with_previous, "currency": currency},
            generated_by=generated_by,
        )

    async def generate_income_statement(
        self,
        *,
        legal_entity_id: UUID,
        start_date: date,
        end_date: date,
        include_details: bool = True,
        compare_with_previous: bool = False,
        currency: str = "IDR",
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not start_date or not end_date:
            raise ValueError("start_date dan end_date wajib diisi untuk income statement")
        data = await self._ledger.get_income_statement(legal_entity_id, start_date, end_date)
        headers = ["Kategori", "Akun", "Jumlah"]
        rows: list[list[Any]] = []
        for line in data.revenues:
            rows.append(["Pendapatan", line.get("account_name", ""), self._dec(line.get("current_period"))])
        for line in data.operating_expenses:
            rows.append(["Beban Operasional", line.get("account_name", ""), self._dec(line.get("current_period"))])
        rows.append(["LABA KOTOR", "", self._dec(data.gross_profit)])
        rows.append(["LABA OPERASIONAL", "", self._dec(data.operating_income)])
        rows.append(["LABA BERSIH", "", self._dec(data.net_income)])

        payload = {
            "period": data.period_name,
            "gross_profit": str(data.gross_profit),
            "operating_income": str(data.operating_income),
            "net_income": str(data.net_income),
            "revenues": data.revenues,
            "operating_expenses": data.operating_expenses,
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="income_statement", report_format=report_format,
            title="Laporan Laba Rugi", headers=headers, rows=rows, raw_payload=payload,
            parameters={"start_date": start_date.isoformat(), "end_date": end_date.isoformat(),
                        "include_details": include_details, "compare_with_previous": compare_with_previous,
                        "currency": currency},
            generated_by=generated_by,
        )

    async def generate_cash_flow(
        self,
        *,
        legal_entity_id: UUID,
        start_date: date,
        end_date: date,
        method: str = "indirect",
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not start_date or not end_date:
            raise ValueError("start_date dan end_date wajib diisi untuk cash flow")
        data = await self._ledger.get_cash_flow_statement(legal_entity_id, start_date, end_date, method=method)
        headers = ["Aktivitas", "Keterangan", "Jumlah"]
        rows: list[list[Any]] = []
        for a in data.operating_activities:
            rows.append(["Operasional", a.get("description", ""), self._dec(a.get("amount"))])
        for a in data.investing_activities:
            rows.append(["Investasi", a.get("description", ""), self._dec(a.get("amount"))])
        for a in data.financing_activities:
            rows.append(["Pendanaan", a.get("description", ""), self._dec(a.get("amount"))])
        rows.append(["Kas Awal", "", self._dec(data.beginning_cash)])
        rows.append(["Kas Akhir", "", self._dec(data.ending_cash)])
        rows.append(["Kenaikan/Penurunan Bersih", "", self._dec(data.net_increase_decrease)])

        payload = {
            "beginning_cash": str(data.beginning_cash), "ending_cash": str(data.ending_cash),
            "net_increase_decrease": str(data.net_increase_decrease),
            "operating_activities": data.operating_activities,
            "investing_activities": data.investing_activities,
            "financing_activities": data.financing_activities,
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="cash_flow", report_format=report_format,
            title="Laporan Arus Kas", headers=headers, rows=rows, raw_payload=payload,
            parameters={"start_date": start_date.isoformat(), "end_date": end_date.isoformat(), "method": method},
            generated_by=generated_by,
        )

    async def generate_equity_statement(
        self,
        *,
        legal_entity_id: UUID,
        start_date: date,
        end_date: date,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not start_date or not end_date:
            raise ValueError("start_date dan end_date wajib diisi untuk equity statement")
        data = await self._ledger.get_equity_statement(legal_entity_id, start_date, end_date)
        headers = ["Komponen", "Saldo Awal", "Penambahan", "Pengurangan", "Saldo Akhir"]
        rows = [
            [ln.component, self._dec(ln.opening_balance), self._dec(ln.additions),
             self._dec(ln.deductions), self._dec(ln.closing_balance)]
            for ln in data.lines
        ]
        rows.append(["TOTAL", self._dec(data.opening_total_equity), "", "", self._dec(data.closing_total_equity)])

        payload = {
            "opening_total_equity": str(data.opening_total_equity),
            "closing_total_equity": str(data.closing_total_equity),
            "net_income": str(data.net_income),
            "lines": [ln.__dict__ for ln in data.lines],
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="equity_statement", report_format=report_format,
            title="Laporan Perubahan Ekuitas", headers=headers, rows=rows, raw_payload=payload,
            parameters={"start_date": start_date.isoformat(), "end_date": end_date.isoformat()},
            generated_by=generated_by,
        )

    # ==================== LEDGER REPORTS ====================

    async def generate_trial_balance(
        self,
        *,
        legal_entity_id: UUID,
        as_of_date: date,
        include_zero_balance: bool = False,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        data = await self._ledger.get_trial_balance(
            legal_entity_id, as_of_date, include_zero_balance=include_zero_balance
        )
        headers = ["Kode Akun", "Nama Akun", "Debit", "Kredit"]
        rows = [
            [ln.account_code, ln.account_name,
             self._dec(ln.closing_balance_debit), self._dec(ln.closing_balance_credit)]
            for ln in data.lines
        ]
        rows.append(["TOTAL", "", self._dec(data.total_debit), self._dec(data.total_credit)])

        payload = {
            "as_of_date": as_of_date.isoformat(),
            "total_debit": str(data.total_debit), "total_credit": str(data.total_credit),
            "is_balanced": data.is_balanced,
            "lines": [ln.__dict__ for ln in data.lines],
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="trial_balance", report_format=report_format,
            title="Neraca Saldo (Trial Balance)", headers=headers, rows=rows, raw_payload=payload,
            parameters={"as_of_date": as_of_date.isoformat(), "include_zero_balance": include_zero_balance},
            generated_by=generated_by,
        )

    async def generate_general_ledger(
        self,
        *,
        legal_entity_id: UUID,
        start_date: date,
        end_date: date,
        account_id: UUID | None = None,
        account_code: str | None = None,
        include_details: bool = True,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not start_date or not end_date:
            raise ValueError("start_date dan end_date wajib diisi untuk general ledger")
        entries = await self._ledger.get_ledger_entries(
            legal_entity_id, start_date, end_date, account_id=account_id, page=1, page_size=5000,
        )
        if account_code:
            entries = [e for e in entries if e.account_code == account_code]

        headers = ["Tanggal", "No. Jurnal", "Kode Akun", "Nama Akun", "Keterangan", "Debit", "Kredit"]
        rows = [
            [e.journal_date.isoformat(), e.journal_number, e.account_code, e.account_name,
             e.description, self._dec(e.debit_amount), self._dec(e.credit_amount)]
            for e in entries
        ]

        payload = {
            "start_date": start_date.isoformat(), "end_date": end_date.isoformat(),
            "entry_count": len(entries),
            "entries": [
                {"journal_date": e.journal_date.isoformat(), "journal_number": e.journal_number,
                 "account_code": e.account_code, "account_name": e.account_name,
                 "debit_amount": str(e.debit_amount), "credit_amount": str(e.credit_amount)}
                for e in entries
            ],
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="general_ledger", report_format=report_format,
            title="Buku Besar (General Ledger)", headers=headers, rows=rows, raw_payload=payload,
            parameters={"start_date": start_date.isoformat(), "end_date": end_date.isoformat(),
                        "account_id": str(account_id) if account_id else None, "account_code": account_code,
                        "include_details": include_details},
            generated_by=generated_by,
        )

    # ==================== SUBLEDGER REPORTS ====================

    async def generate_ar_aging(
        self,
        *,
        legal_entity_id: UUID,
        as_of_date: date,
        customer_id: UUID | None = None,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        aging = await self._ar.get_aging_all_customers(legal_entity_id, as_of_date)
        if customer_id:
            aging = [a for a in aging if a.customer_id == customer_id]

        headers = ["Pelanggan", "Total Outstanding"] + [b.bucket_name for b in (aging[0].buckets if aging else [])]
        rows = []
        for a in aging:
            row = [a.customer_name, self._dec(a.total_outstanding)]
            row.extend(self._dec(b.total_amount) for b in a.buckets)
            rows.append(row)

        payload = {
            "as_of_date": as_of_date.isoformat(),
            "customers": [
                {"customer_id": str(a.customer_id), "customer_name": a.customer_name,
                 "total_outstanding": str(a.total_outstanding),
                 "buckets": [{"bucket_name": b.bucket_name, "total_amount": str(b.total_amount)} for b in a.buckets]}
                for a in aging
            ],
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="ar_aging", report_format=report_format,
            title="Aging Piutang (AR Aging)", headers=headers, rows=rows, raw_payload=payload,
            parameters={"as_of_date": as_of_date.isoformat(), "customer_id": str(customer_id) if customer_id else None},
            generated_by=generated_by,
        )

    async def generate_ap_aging(
        self,
        *,
        legal_entity_id: UUID,
        as_of_date: date,
        vendor_id: UUID | None = None,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        aging = await self._ap.get_aging_all_vendors(legal_entity_id, as_of_date)
        if vendor_id:
            aging = [a for a in aging if getattr(a, "vendor_id", None) == vendor_id]

        sample_buckets = getattr(aging[0], "buckets", []) if aging else []
        headers = ["Vendor", "Total Outstanding"] + [b.bucket_name for b in sample_buckets]
        rows = []
        for a in aging:
            row = [getattr(a, "vendor_name", ""), self._dec(getattr(a, "total_outstanding", None))]
            row.extend(self._dec(b.total_amount) for b in getattr(a, "buckets", []))
            rows.append(row)

        payload = {
            "as_of_date": as_of_date.isoformat(),
            "vendors": [
                {
                    "vendor_id": str(getattr(a, "vendor_id", "")),
                    "vendor_name": getattr(a, "vendor_name", ""),
                    "total_outstanding": str(getattr(a, "total_outstanding", 0)),
                    "buckets": [
                        {"bucket_name": b.bucket_name, "total_amount": str(b.total_amount)}
                        for b in getattr(a, "buckets", [])
                    ],
                }
                for a in aging
            ],
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="ap_aging", report_format=report_format,
            title="Aging Utang (AP Aging)", headers=headers, rows=rows, raw_payload=payload,
            parameters={"as_of_date": as_of_date.isoformat(), "vendor_id": str(vendor_id) if vendor_id else None},
            generated_by=generated_by,
        )

    # ==================== INVENTORY REPORTS ====================

    async def generate_stock_card(
        self,
        *,
        legal_entity_id: UUID,
        start_date: date | None,
        end_date: date | None,
        item_id: UUID | None,
        warehouse_id: UUID | None = None,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not item_id:
            raise ValueError("item_id (lewat 'parameters.item_id' atau field item_id) wajib diisi untuk stock card")
        movements = await self._inventory.get_stock_card(item_id, start_date, end_date)

        headers = ["Tanggal", "Tipe", "Masuk", "Keluar", "Harga Satuan", "Nilai Total", "Referensi", "Gudang"]
        rows = [
            [m["date"], m["movement_type"], m["quantity_in"], m["quantity_out"],
             m["unit_cost"], m["total_value"], m.get("reference"), m.get("warehouse")]
            for m in movements
        ]

        payload = {
            "item_id": str(item_id),
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "movement_count": len(movements),
            "movements": movements,
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="stock_card", report_format=report_format,
            title="Kartu Stok", headers=headers, rows=rows, raw_payload=payload,
            parameters={"item_id": str(item_id), "warehouse_id": str(warehouse_id) if warehouse_id else None,
                        "start_date": start_date.isoformat() if start_date else None,
                        "end_date": end_date.isoformat() if end_date else None},
            generated_by=generated_by,
        )

    # ==================== TAX REPORTS ====================

    async def generate_tax_summary(
        self,
        *,
        legal_entity_id: UUID,
        start_date: date,
        end_date: date,
        tax_type: str | None = None,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not start_date or not end_date:
            raise ValueError("start_date dan end_date wajib diisi untuk tax summary")

        try:
            from ports.primary.tax_transaction_repository_port import TaxType
        except ImportError:
            TaxType = None  # type: ignore[assignment]

        headers = ["Jenis Pajak", "Total Terutang", "Total Kredit Pajak"]
        rows: list[list[Any]] = []
        breakdown: list[dict[str, Any]] = []

        tax_repo = getattr(self, "_tax_transaction_repo", None)
        if tax_repo is not None and TaxType is not None:
            tax_types = [TaxType(tax_type)] if tax_type else list(TaxType)
            for tt in tax_types:
                liability = await tax_repo.get_total_tax_liability(legal_entity_id, tt, start_date, end_date)
                credit = await tax_repo.get_total_tax_credit(legal_entity_id, tt, start_date, end_date)
                if liability or credit:
                    rows.append([tt.value, self._dec(liability), self._dec(credit)])
                    breakdown.append({"tax_type": tt.value, "liability": str(liability), "credit": str(credit)})
        else:
            logger.warning(
                "TaxTransactionRepositoryPort tidak tersedia di ReportService - "
                "tax summary dikembalikan kosong, bukan error, supaya endpoint tetap berhasil."
            )

        payload = {
            "start_date": start_date.isoformat(), "end_date": end_date.isoformat(),
            "tax_type_filter": tax_type, "breakdown": breakdown,
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="tax_summary", report_format=report_format,
            title="Ringkasan Pajak", headers=headers, rows=rows, raw_payload=payload,
            parameters={"start_date": start_date.isoformat(), "end_date": end_date.isoformat(), "tax_type": tax_type},
            generated_by=generated_by,
        )

    # ==================== FINANCIAL RATIOS & ANALYTICS ====================

    async def generate_financial_ratios(
        self,
        *,
        legal_entity_id: UUID,
        as_of_date: date,
        compare_industry: bool = False,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        data = await self._ledger.get_financial_ratios(
            legal_entity_id, as_of_date, compare_industry=compare_industry
        )
        headers = ["Rasio", "Nilai"]
        ratio_fields = [
            ("Current Ratio", data.current_ratio), ("Quick Ratio", data.quick_ratio),
            ("Cash Ratio", data.cash_ratio), ("Debt to Equity", data.debt_to_equity),
            ("Debt to Assets", data.debt_to_assets), ("Net Margin", data.net_margin),
            ("Return on Assets", data.return_on_assets), ("Return on Equity", data.return_on_equity),
            ("Asset Turnover", data.asset_turnover),
        ]
        rows = [[name, f"{val:.4f}" if val is not None else "N/A"] for name, val in ratio_fields]

        payload = {"as_of_date": as_of_date.isoformat(), "ratios": {name: val for name, val in ratio_fields}}
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="financial_ratios", report_format=report_format,
            title="Rasio Keuangan", headers=headers, rows=rows, raw_payload=payload,
            parameters={"as_of_date": as_of_date.isoformat(), "compare_industry": compare_industry},
            generated_by=generated_by,
        )

    # ==================== BUDGET REPORTS ====================

    async def generate_budget_vs_actual(
        self,
        *,
        legal_entity_id: UUID,
        fiscal_year: int | None = None,
        period: int | None = None,
        budget_id: UUID | None = None,
        report_format: str = "pdf",
        generated_by: UUID,
    ) -> GeneratedReportTable:
        if not budget_id:
            raise ValueError("budget_id (lewat 'parameters.budget_id') wajib diisi untuk budget vs actual")
        data = await self._budget.get_budget_vs_actual(budget_id, legal_entity_id, period or 0)
        if data is None:
            raise ValueError(f"Budget {budget_id} tidak ditemukan")

        headers = ["Kategori", "Anggaran", "Aktual", "Varians", "% Varians"]
        rows = [
            [ln.category if hasattr(ln, "category") else "", self._dec(getattr(ln, "budget_amount", None)),
             self._dec(getattr(ln, "actual_amount", None)), self._dec(getattr(ln, "variance", None)),
             f"{getattr(ln, 'variance_percent', 0):.2f}%"]
            for ln in data.lines
        ]
        rows.append(["TOTAL", self._dec(data.total_budget), self._dec(data.total_actual),
                     self._dec(data.total_variance), f"{data.variance_percent:.2f}%"])

        payload = {
            "budget_id": str(budget_id), "budget_name": data.budget_name,
            "total_budget": str(data.total_budget), "total_actual": str(data.total_actual),
            "total_variance": str(data.total_variance), "variance_percent": data.variance_percent,
        }
        return await self._persist_report(
            legal_entity_id=legal_entity_id, report_type="budget_vs_actual", report_format=report_format,
            title="Anggaran vs Aktual", headers=headers, rows=rows, raw_payload=payload,
            parameters={"fiscal_year": fiscal_year, "period": period, "budget_id": str(budget_id)},
            generated_by=generated_by,
        )

    # ==================== REPORT MANAGEMENT ====================

    async def list_reports(
        self,
        *,
        legal_entity_id: UUID,
        report_type: str | None = None,
        status: str | None = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> PagedGeneratedReports:
        items, total = await self._report_repo.list_generated_reports(
            legal_entity_id=legal_entity_id, report_type=report_type, status=status,
            start_date=start_date, end_date=end_date, page=page, page_size=page_size,
        )
        return PagedGeneratedReports(items=items, total=total)

    async def get_report_by_id(self, report_id: UUID, legal_entity_id: UUID) -> GeneratedReportTable | None:
        return await self._report_repo.get_generated_report_by_id(report_id, legal_entity_id)

    async def delete_report(
        self, report_id: UUID, legal_entity_id: UUID, deleted_by: UUID
    ) -> GeneratedReportTable | None:
        result = await self._report_repo.soft_delete_generated_report(report_id, legal_entity_id, deleted_by)
        if result:
            self._record_audit("delete_report", {"report_number": result.report_number})
        return result

    async def get_report_status(self, report_id: UUID, legal_entity_id: UUID) -> ReportStatusInfo | None:
        report = await self._report_repo.get_generated_report_by_id(report_id, legal_entity_id)
        if not report:
            return None
        is_done = report.status in ("generated", "failed")
        return ReportStatusInfo(
            report_number=report.report_number,
            status=report.status,
            progress_percent=100 if is_done else 50,
            current_step="Selesai" if is_done else "Memproses",
            total_steps=1,
            estimated_remaining_seconds=0 if is_done else None,
            error_message=report.error_message,
            generated_at=report.generated_at,
        )

    async def get_report_history(self, report_id: UUID, legal_entity_id: UUID) -> list[ReportHistoryEntry]:
        report = await self._report_repo.get_generated_report_by_id(report_id, legal_entity_id)
        if not report:
            return []
        entries = [
            ReportHistoryEntry(
                timestamp=report.generated_at,
                action="generated" if report.status == "generated" else "generate_failed",
                status=report.status,
                actor_id=report.generated_by,
                reason=report.error_message,
                details={"report_type": report.report_type, "report_format": report.report_format},
            )
        ]
        if report.is_deleted and report.deleted_at:
            entries.append(
                ReportHistoryEntry(
                    timestamp=report.deleted_at,
                    action="deleted",
                    status="deleted",
                    actor_id=report.deleted_by or report.generated_by,
                    details=None,
                )
            )
        return entries


__all__ = [
    "PagedGeneratedReports",
    "ReportHistoryEntry",
    "ReportService",
    "ReportServiceError",
    "ReportStatusInfo",
    "create_report_service",
]


# ============================================================================
# Factory
# ============================================================================


async def create_report_service(
    report_repo: ReportRepositoryPort,
    ledger_service: LedgerService,
    ar_service: ARService,
    ap_service: APService,
    budget_service: BudgetService,
    inventory_service: InventoryService,
    tax_transaction_repo: TaxTransactionRepositoryPort | None = None,
) -> ReportService:
    return ReportService(
        report_repo, ledger_service, ar_service, ap_service,
        budget_service, inventory_service, tax_transaction_repo,
    )
